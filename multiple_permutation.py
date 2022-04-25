from rail_fence import encrypt_rail, dencrypt_rail
from simple_permutation_cipher import encrypt_simple, dencrypt_simple
from vertical_permutation_cipher import encrypt_vertical, dencrypt_vertical
import random
import re
import string
import openpyxl
from openpyxl import load_workbook

def save_cipher(name,name_cipher,key,size,str1):
    fn = "cipher.xlsx"
    wb = load_workbook(filename="cipher.xlsx", data_only=True)
    ws = wb['Sheet1']
    flag=0
    for x in range(1, ws.max_row + 1):
        if ws.cell(row=x, column=1).value==name:
            ws['B'+str(x)].value = name_cipher
            ws['C'+str(x)] = str(key)
            ws['D'+str(x)] = str1
            ws['E'+str(x)] = str(size)
            flag=1
    if flag==0:
        ws.append([name, name_cipher, key, str1, size])
    wb.save(fn)
    wb.close()

def generate_random_string(length):
    letters = string.ascii_lowercase
    rand_string = ''.join(random.choice(letters) for i in range(length))
    return rand_string

def encrypt_multiple(text,choice_00):
    lst=list(range(1, 5, 1))
    random.shuffle(lst)
    for i in lst:
        if i == 4:
            i = random.randint(1, 4)
        if i==1:
            n = random.randint(1, 10)
            key = [str(i) for i in range(0, n, 1)]
            random.shuffle(key)
            key = ''.join(key)
            size=random.randint(1, 5)
            name_cipher = "Шифр простой перестановки"
            print(name_cipher,",","Ключ:",key, " ,","Количество символов", size)
            if choice_00 == '1':
                str1='Без добавления нулей'
            else:
                str1 = 'С добавлением нулей'
            text=encrypt_simple(key,text,size)
            name = generate_random_string(5)
            save_cipher(name, name_cipher, key, size, str1)
            if choice_00 == '1':
                s=text.replace('\0','')
                if '\a' in text:
                    s=s.replace('\a','')
                print(s)
            else:
                print(text)
        if i==2:
            m = str(random.randint(2, 9))
            n = random.randint(2, 9)
            k = str(n)
            help = [str(i) for i in range(0, n, 1)]
            random.shuffle(help)
            help = ''.join(help)
            key="{0}*{1} {2}".format(m,k,help)
            size = random.randint(1, 5)
            name_cipher="Шифр вертикальной перестановки"
            print(name_cipher, ",", "Ключ:", key, " ,", "Количество символов", size)
            if choice_00 == '1':
                str1='Без добавления нулей'
            else:
                str1 = 'С добавлением нулей'
            name=generate_random_string(6)
            save_cipher(name,name_cipher,key,size,str1)
            text = encrypt_vertical(key, text, size)
            if choice_00 == '1':
                s = text.replace('\a', '')
                s = s.replace('\0', '')
                print(s)
            else:
                print(text)
        if i==3:
            m = random.randint(3, 10)
            n = random.randint(3, 10)
            size = random.randint(1, 5)
            text = encrypt_rail(m,n, text, size)
            if choice_00 == '1':
                str1='Без добавления нулей'
            else:
                str1 = 'С добавлением нулей'
            name=generate_random_string(4)
            key="{}*{}".format(m,n)
            name_cipher = "Шифр Rail Fence"
            print(name_cipher, ",", "Размер матрицы:", key, " ,", "Количество символов", size)
            save_cipher(name,name_cipher,key,size,str1)
            if choice_00 == '1':
                s = text.replace('\0', '')
                if '\a' in text:
                    s=s.replace('\a','')
                print(s)
            else:
                print(text)
        i+=1
