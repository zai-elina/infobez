import random
from rail_fence import encrypt_rail, dencrypt_rail
from simple_permutation_cipher import encrypt_simple, dencrypt_simple
from vertical_permutation_cipher import encrypt_vertical, dencrypt_vertical
from multiple_permutation import encrypt_multiple
import openpyxl
from openpyxl import load_workbook

def choosing_a_simple_permutation():
    choice_of_key = input('Выберите способ задания ключа:\n 1)Ввести свой ключ\n 2)Случайный ключ\n')
    if choice_of_key == "1":
        key = input('Введите ключ\n')
    elif choice_of_key == '2':
        n = random.randint(3, 10)
        key = [str(i) for i in range(0, n, 1)]
        random.shuffle(key)
        key = ''.join(key)
        print('Получившийся ключ:', key)
    choice_0=input('Выберите способ шифрования последнего блока исходного текста, размер которого меньше размера блока шифрования:\n 1)Без добавления нулей\n 2)С добавлением нулей\n')
    choice_size=input('Выберите элементы блока:\n1)отдельный символ\n2)группа из произвольного количества символов\n')
    if choice_size=='1':
        size=1
    elif choice_size=='2':
        size=int(input('Введите количеств символов в блоке\n'))
    return key, choice_0,size

def choosing_a_vertical_permutation():
    choice_of_key = input('Выберите способ задания ключа:\n 1)Ввести свой ключ\n 2)Случайный ключ\n')
    if choice_of_key == "1":
        key = input('Введите ключ в следующем виде "2*4 1302", где сначала указывается размер матрицы, затем сам ключ\n Учтите, что количество столбцов равно длине ключа\n')
    elif choice_of_key == '2':
        m = str(random.randint(2, 9))
        n = random.randint(2, 9)
        k = str(n)
        help = [str(i) for i in range(0, n, 1)]
        random.shuffle(help)
        help = ''.join(help)
        key = "{0}*{1} {2}".format(m, k, help)
        print('Получившийся ключ:', key)
    choice_0 = input(
        'Выберите способ шифрования последнего блока исходного текста, размер которого меньше размера блока шифрования:\n 1)Без добавления нулей\n 2)С добавлением нулей\n')
    choice_size = input(
        'Выберите элементы блока:\n1)отдельный символ\n2)группа из произвольного количества символов\n')
    if choice_size == '1':
        size = 1
    elif choice_size == '2':
        size = int(input('Введите количеств символов в блоке\n'))
    return key, choice_0, size

def choosing_a_rail_fence():
    choice_of_key = input('Выберите способ задания ключа:\n 1)Ввести свой ключ\n 2)Случайный ключ\n')
    if choice_of_key == "1":
        m = int(input(
            'Введите количество строк\n'))
        n = int(input(
            'Введите количество столбцов(>3)\n'))
    elif choice_of_key == '2':
        m = random.randint(3,5)
        n = random.randint(6, 10)
        print('Получившийся резмер:', '{}*{}'.format(m,n))
    choice_0 = input(
        'Выберите способ шифрования последнего блока исходного текста, размер которого меньше размера блока шифрования:\n 1)Без добавления нулей\n 2)С добавлением нулей\n')
    choice_size = input(
        'Выберите элементы блока:\n1)отдельный символ\n2)группа из произвольного количества символов\n')
    if choice_size == '1':
        size = 1
    elif choice_size == '2':
        size = int(input('Введите количеств символов в блоке\n'))
    return m,n, choice_0, size

def save_cipher(name,name_cipher,key,size,str1):
    fn = "cipher.xlsx"
    wb = load_workbook(filename="cipher.xlsx", data_only=True)
    ws = wb['Sheet1']
    flag=0
    for x in range(1, ws.max_row + 1):
        if ws.cell(row=x, column=1).value==name:
            ws['B'+str(x)].value = name_cipher
            ws['C'+str(x)] = str(key)
            ws['D'+str(x)] = str1
            ws['E'+str(x)] = str(size)
            flag=1
    if flag==0:
        ws.append([name, name_cipher, key, str1, size])
    wb.save(fn)
    wb.close()

def print_inf(name):
    fn = "cipher.xlsx"
    wb = load_workbook(filename="cipher.xlsx", data_only=True)
    ws = wb['Sheet1']
    flag=0
    for x in range(1, ws.max_row + 1):
        if ws.cell(row=x, column=1).value==name:
            print("Название шифра:",ws.cell(row=x, column=2).value,",", "Ключ шифрования:",
            ws.cell(row=x, column=3).value, ",",
            ws.cell(row=x, column=4).value,",","Количество символов",
            ws.cell(row=x, column=5).value)
            flag=1
    if flag==0:
        print("Ключа с таким названием не существует")
    wb.save(fn)
    wb.close()

if __name__ == "__main__":
    while True:
        print('Выберите алгоритм шифрования:')
        print('1)Шифр простой перестановки')
        print('2)Шифр вертикальной перестановки')
        print('3)Шифр Rail Fence ')
        print('4)Шифр множественной перестановки')
        print('5)Вывести информацию о ключе')
        print('6)Выйти')
        text_vybor = input()
        if text_vybor == "1":
            key, choice_0, size=choosing_a_simple_permutation()
            choice_text=input('Выберите способ ввода текста:\n 1)Через консоль\n 2)Из файла\n')
            if choice_text=='1':
                text=input('Введите текст\n')
            elif choice_text=='2':
                file=input("Введите имя файла:  ")
                with open(file, encoding='utf-8') as f:
                    text = f.read()
            text=encrypt_simple(key,text,size)
            if choice_0 == '1':
                str1='Без добавления нулей'
            else:
                str1 = 'С добавлением нулей'
            name=input("Название шифра   ")
            name_cipher="Шифр простой перестановки"
            save_cipher(name,name_cipher,key,size,str1)
            if choice_0=='1':
                text1 = text.replace('\0', '')
                sms_out=input("Куда вы хотите вывести зашифрованное сообщение:\n1)В консоль\n2)В файл\n")
                if sms_out=='1':
                    print("Зашифрованное сообщение:", text1 + '\n')
                elif sms_out=='2':
                    with open("encrypted1_text.txt", 'w', encoding='utf-8') as f:
                        f.write(text1)
                        print("Зашифрованный текст в файле encrypted1_text.txt.")
            else:
                sms_out = input("Куда вы хотите вывести зашифрованное сообщение:\n1)В консоль\n2)В файл\n")
                if sms_out == '1':
                    print("Зашифрованное сообщение:", text + '\n')
                elif sms_out == '2':
                    with open("encrypted1_text.txt", 'w', encoding='utf-8') as f:
                        f.write(text)
                        print("Зашифрованный текст в файле encrypted1_text.txt.")
            dec=input("Вы хотите расшифровать?(Да/Нет)\n")
            if dec=='Да':
                choice_text = input('Выберите способ ввода текста:\n 1)Введеное ранее сообщение\n 2)Из файла\n')
                if choice_text == '1':
                    text = text
                elif choice_text == '2':
                    file = input("Введите имя файла:  ")
                    with open(file, encoding='utf-8') as f:
                        text = f.read()
                text = dencrypt_simple(key, text, size)
                sms_out1 = input("Куда вы хотите вывести расшифрованное сообщение:\n1)В консоль\n2)В файл\n")
                if sms_out1 == '1':
                    print("Расшифрованное текст:", text + '\n')
                elif sms_out1 == '2':
                    with open("normal1_text.txt", 'w', encoding='utf-8') as f:
                        f.write(text)
                        print("Расшифрованное текст в файле normal1_text.txt.")
        elif text_vybor == "2":
            key, choice_0, size = choosing_a_vertical_permutation()
            choice_text = input('Выберите способ ввода текста:\n 1)Через консоль\n 2)Из файла\n')
            if choice_text == '1':
                text = input('Введите текст\n')
            elif choice_text == '2':
                file = input("Введите имя файла:  ")
                with open(file, encoding='utf-8') as f:
                    text = f.read()
            text = encrypt_vertical(key, text, size)
            if choice_0 == '1':
                str1='Без добавления нулей'
            else:
                str1 = 'С добавлением нулей'
            name=input("Название шифра  ")
            name_cipher="Шифр вертикальной перестановки"
            save_cipher(name,name_cipher,key,size,str1)
            if choice_0 == '1':
                text1 = text.replace('\a', '')
                sms_out = input("Куда вы хотите вывести зашифрованное сообщение:\n1)В консоль\n2)В файл\n")
                if sms_out == '1':
                    print("Зашифрованное сообщение:", text1 + '\n')
                elif sms_out == '2':
                    with open("encrypted2_text.txt", 'w', encoding='utf-8') as f:
                        f.write(text1)
                        print("Зашифрованный текст в файле encrypted2_text.txt.")
            else:
                sms_out = input("Куда вы хотите вывести зашифрованное сообщение:\n1)В консоль\n2)В файл\n")
                if sms_out == '1':
                    print("Зашифрованное сообщение:", text + '\n')
                elif sms_out == '2':
                    with open("encrypted2_text.txt", 'w', encoding='utf-8') as f:
                        f.write(text)
                        print("Зашифрованный текст в файле encrypted2_text.txt.")
            dec = input("Вы хотите расшифровать?(Да/Нет)\n")
            if dec == 'Да':
                choice_text = input('Выберите способ ввода текста:\n 1)Введеное ранее сообщение\n 2)Из файла\n')
                if choice_text == '1':
                    text = text
                elif choice_text == '2':
                    file = input("Введите имя файла:  ")
                    with open(file, encoding='utf-8') as f:
                        text = f.read()
                text = dencrypt_vertical(key, text, size)
                sms_out1 = input("Куда вы хотите вывести расшифрованное сообщение:\n1)В консоль\n2)В файл\n")
                if sms_out1 == '1':
                    print("Расшифрованное текст:", text + '\n')
                elif sms_out1 == '2':
                    with open("normal2_text.txt", 'w', encoding='utf-8') as f:
                        f.write(text)
                        print("Расшифрованное текст в файле normal2_text.txt.")
        elif text_vybor == "3":
            m,n, choice_0, size = choosing_a_rail_fence()
            choice_text = input('Выберите способ ввода текста:\n 1)Через консоль\n 2)Из файла\n')
            if choice_text == '1':
                text = input('Введите текст\n')
            elif choice_text == '2':
                file = input("Введите имя файла:  ")
                with open(file, encoding='utf-8') as f:
                    text = f.read()
            text = encrypt_rail(m,n, text, size)
            if choice_0 == '1':
                str1='Без добавления нулей'
            else:
                str1 = 'С добавлением нулей'
            name=input("Название шифра  ")
            name_cipher="Шифр Rail Fence"
            key="{}*{}".format(m,n)
            save_cipher(name,name_cipher,key,size,str1)
            if choice_0 == '1':
                text1 = text.replace('\0', '')
                sms_out = input("Куда вы хотите вывести зашифрованное сообщение:\n1)В консоль\n2)В файл\n")
                if sms_out == '1':
                    print("Зашифрованное сообщение:", text1 + '\n')
                elif sms_out == '2':
                    with open("encrypted3_text.txt", 'w', encoding='utf-8') as f:
                        f.write(text1)
                        print("Зашифрованный текст в файле encrypted3_text.txt.")
            else:
                sms_out = input("Куда вы хотите вывести зашифрованное сообщение:\n1)В консоль\n2)В файл\n")
                if sms_out == '1':
                    print("Зашифрованное сообщение:", text + '\n')
                elif sms_out == '2':
                    with open("encrypted3_text.txt", 'w', encoding='utf-8') as f:
                        f.write(text)
                        print("Зашифрованный текст в файле encrypted3_text.txt.")
            dec = input("Вы хотите расшифровать?(Да/Нет)\n")
            if dec == 'Да':
                choice_text = input('Выберите способ ввода текста:\n 1)Введеное ранее сообщение\n 2)Из файла\n')
                if choice_text == '1':
                    text = text
                elif choice_text == '2':
                    file = input("Введите имя файла:  ")
                    with open(file, encoding='utf-8') as f:
                        text = f.read()
                text = dencrypt_rail(m,n, text, size)
                sms_out1 = input("Куда вы хотите вывести расшифрованное сообщение:\n1)В консоль\n2)В файл\n")
                if sms_out1 == '1':
                    print("Расшифрованное текст:", text + '\n')
                elif sms_out1 == '2':
                    with open("normal3_text.txt", 'w', encoding='utf-8') as f:
                        f.write(text)
                        print("Расшифрованное текст в файле normal3_text.txt.")
        elif text_vybor == "4":
            choice_00 = input(
            'Выберите способ шифрования последнего блока исходного текста, размер которого меньше размера блока шифрования:\n 1)Без добавления нулей\n 2)С добавлением нулей\n' + '\n')
            choice_text = input('Выберите способ ввода текста:\n 1)Через консоль\n 2)Из файла\n')
            if choice_text == '1':
                text = input('Введите текст\n')
            elif choice_text == '2':
                file = input("Введите имя файла:  ")
                with open(file, encoding='utf-8') as f:
                    text = f.read()
            encrypt_multiple(text,choice_00)
        elif text_vybor == "5":
            name=input('Название  ')
            print_inf(name)
        elif text_vybor == "6":
            break
        else:
            print('Введите 1,2,3,4,5 или 6')
